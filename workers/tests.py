from django.test import TestCase
from django.contrib.auth.models import User
from django.urls import reverse
from rest_framework.test import APITestCase
from rest_framework import status
from django.core.files.uploadedfile import SimpleUploadedFile
import io
import openpyxl
from .models import Worker


class WorkerModelTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="testuser", email="test@example.com", password="testpass123"
        )

    def test_worker_creation(self):
        """Тест создания работника"""
        worker = Worker.objects.create(
            first_name="Иван",
            last_name="Петров",
            email="ivan@test.com",
            position="Разработчик",
            created_by=self.user,
        )
        self.assertEqual(worker.first_name, "Иван")
        self.assertEqual(worker.last_name, "Петров")
        self.assertEqual(worker.email, "ivan@test.com")
        self.assertTrue(worker.is_active)
        self.assertFalse(worker.is_deleted)

    def test_worker_str_method(self):
        """Тест строкового представления работника"""
        worker = Worker.objects.create(
            first_name="Иван",
            last_name="Петров",
            email="ivan@test.com",
            position="Разработчик",
            created_by=self.user,
        )
        expected = f"{worker.first_name} {worker.last_name}"
        self.assertEqual(str(worker), expected)


class WorkerAPITest(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="testuser", email="test@example.com", password="testpass123"
        )
        self.admin_user = User.objects.create_user(
            username="admin",
            email="admin@example.com",
            password="adminpass123",
            is_staff=True,
        )
        self.worker = Worker.objects.create(
            first_name="Иван",
            last_name="Петров",
            email="ivan@test.com",
            position="Разработчик",
            created_by=self.user,
        )

    def test_get_workers_list(self):
        """Тест получения списка работников"""
        url = reverse("worker-list")
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data["results"]), 1)

    def test_get_worker_detail(self):
        """Тест получения детальной информации о работнике"""
        url = reverse("worker-detail", kwargs={"pk": self.worker.pk})
        response = self.client.get(url)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["first_name"], "Иван")

    def test_create_worker_as_admin(self):
        """Тест создания работника администратором"""
        self.client.force_authenticate(user=self.admin_user)
        url = reverse("worker-list")
        data = {
            "first_name": "Мария",
            "last_name": "Сидорова",
            "email": "maria@test.com",
            "position": "Дизайнер",
        }
        response = self.client.post(url, data)
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(Worker.objects.count(), 2)

    def test_create_worker_as_regular_user(self):
        """Тест создания работника обычным пользователем (должно быть запрещено)"""
        self.client.force_authenticate(user=self.user)
        url = reverse("worker-list")
        data = {
            "first_name": "Мария",
            "last_name": "Сидорова",
            "email": "maria@test.com",
            "position": "Дизайнер",
        }
        response = self.client.post(url, data)
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_update_worker(self):
        """Тест обновления работника"""
        self.client.force_authenticate(user=self.admin_user)
        url = reverse("worker-detail", kwargs={"pk": self.worker.pk})
        data = {"position": "Старший разработчик"}
        response = self.client.patch(url, data)
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.worker.refresh_from_db()
        self.assertEqual(self.worker.position, "Старший разработчик")

    def test_delete_worker(self):
        """Тест удаления работника (мягкое удаление)"""
        self.client.force_authenticate(user=self.admin_user)
        url = reverse("worker-detail", kwargs={"pk": self.worker.pk})
        response = self.client.delete(url)
        self.assertEqual(response.status_code, status.HTTP_204_NO_CONTENT)
        self.worker.refresh_from_db()
        self.assertTrue(self.worker.is_deleted)

    def test_filter_by_is_active(self):
        """Тест фильтрации по активности"""
        Worker.objects.create(
            first_name="Анна",
            last_name="Козлова",
            email="anna@test.com",
            position="Тестировщик",
            is_active=False,
            created_by=self.user,
        )
        url = reverse("worker-list")
        response = self.client.get(url, {"is_active": "true"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data["results"]), 1)

    def test_filter_by_position(self):
        """Тест фильтрации по должности"""
        url = reverse("worker-list")
        response = self.client.get(url, {"position": "Разработчик"})
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data["results"]), 1)


class WorkerImportTest(APITestCase):
    def setUp(self):
        self.admin_user = User.objects.create_user(
            username="admin",
            email="admin@example.com",
            password="adminpass123",
            is_staff=True,
        )

    def create_excel_file(self, data):
        """Создает Excel файл для тестирования"""
        wb = openpyxl.Workbook()
        ws = wb.active

        # Заголовки
        headers = [
            "first_name",
            "middle_name",
            "last_name",
            "email",
            "position",
            "is_active",
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Данные
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Сохраняем в память
        file_content = io.BytesIO()
        wb.save(file_content)
        file_content.seek(0)

        return SimpleUploadedFile(
            "test_workers.xlsx",
            file_content.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_import_workers_success(self):
        """Тест успешного импорта работников"""
        self.client.force_authenticate(user=self.admin_user)

        data = [
            ["Иван", "Петрович", "Иванов", "ivan@import.com", "Разработчик", True],
            ["Мария", "", "Сидорова", "maria@import.com", "Дизайнер", False],
        ]

        excel_file = self.create_excel_file(data)
        url = "/api/workers/import/"
        response = self.client.post(url, {"file": excel_file}, format="multipart")

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["Создан"], 2)
        self.assertEqual(len(response.data["Ошибки"]), 0)
        self.assertEqual(Worker.objects.count(), 2)

    def test_import_workers_with_errors(self):
        """Тест импорта с ошибками валидации"""
        self.client.force_authenticate(user=self.admin_user)

        data = [
            ["", "", "Иванов", "ivan@import.com", "Разработчик", True],  # Нет имени
            ["Мария", "", "", "maria@import.com", "Дизайнер", False],  # Нет фамилии
        ]

        excel_file = self.create_excel_file(data)
        url = "/api/workers/import/"
        response = self.client.post(url, {"file": excel_file}, format="multipart")

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["Создан"], 0)
        self.assertEqual(len(response.data["Ошибки"]), 2)

    def test_import_without_file(self):
        """Тест импорта без файла"""
        self.client.force_authenticate(user=self.admin_user)
        url = "/api/workers/import/"
        response = self.client.post(url, {})

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("Ошибка", response.data)
