from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, SAFE_METHODS, BasePermission
from django_filters.rest_framework import DjangoFilterBackend
from .models import Worker
from .serializers import WorkerSerializer, WorkerDetailSerializer, FileUploadSerializer
import openpyxl
import io
from django.db import IntegrityError


class IsAdminOrReadOnly(BasePermission):
    def has_permission(self, request, view):
        if request.method in SAFE_METHODS:
            return True
        return request.user and request.user.is_staff
    
class WorkerViewSet(viewsets.ModelViewSet):
    queryset = Worker.objects.filter(is_deleted=False).select_related('created_by')
    serializer_class = WorkerDetailSerializer
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['is_active', 'position']

    def get_serializer_class(self):
        if self.action == 'list':
            return WorkerSerializer
        elif self.action == 'import_workers':
            return FileUploadSerializer
        return WorkerDetailSerializer
    
    def perform_create(self, serializer):
        isinstance = serializer.save(created_by=self.request.user  )
        print(f"[LOG] Работник создан: {isinstance} by {self.request.user}")
    
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.is_deleted = True
        instance.save()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['get', 'post'], url_path='import')
    def import_workers(self, request):
        if request.method == 'GET':
            return Response({'message': 'Загрузите Excel файл для импорта сотрудников'})
        
        file = request.FILES.get('file')
        if not file:
            return Response({'Ошибка': 'файл не найден или не существует'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Читаем файл в память
            file_content = file.read()
            file_stream = io.BytesIO(file_content)
            wb = openpyxl.load_workbook(file_stream)
        except Exception as e:
            return Response({'Ошибка': f'Не удалось прочитать Excel файл: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        created = 0
        errors = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            data = dict(zip(headers, row))
            if not data.get('first_name') \
                    or not data.get('last_name') \
                    or not data.get('email') \
                    or not data.get('position'):
                errors.append(f"Строка {idx}: отсутствуют обязательные поля")
                continue
            try:
                worker = Worker.objects.create(
                    first_name=data.get('first_name'),
                    middle_name=data.get('middle_name'),
                    last_name=data.get('last_name'),
                    email=data.get('email'),
                    position=data.get('position'),
                    is_active=bool(data.get('is_active')) if data.get('is_active') is not None else True,
                    created_by=request.user if request.user.is_authenticated else None
                )
                created += 1
            except IntegrityError as e:
                errors.append(f"Строка {idx}: Ошибка: {str(e)}")
            except Exception as e:
                errors.append(f"Строка {idx}: Ошибка: {str(e)}")    
        return Response({'Создан': created, 'Ошибки': errors}, status=status.HTTP_201_CREATED)